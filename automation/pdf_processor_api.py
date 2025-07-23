import os
import io
import base64
import json
import uuid
import threading
import time
from datetime import datetime, timedelta
from queue import Queue
from flask import Flask, request, jsonify, send_file, Response
from flask_cors import CORS
from google.cloud import vision
import pandas as pd
import numpy as np
from PIL import Image
import pdf2image
import tempfile
import re
from typing import Dict, List
from dotenv import load_dotenv
import openpyxl
from openpyxl.utils import get_column_letter
from openai import OpenAI
import httpx

# Load environment variables
load_dotenv('config.env') # Main config
load_dotenv() # Load .env for OPENAI_API_KEY

# Initialize OpenAI client
openai_api_key = os.getenv('OPENAI_API_KEY')
if not openai_api_key:
    print("Warning: OPENAI_API_KEY is not set. OpenAI features will be disabled.")
    openai_client = None
else:
    # Explicitly create an httpx client
    http_client = httpx.Client()
    openai_client = OpenAI(api_key=openai_api_key, http_client=http_client)

app = Flask(__name__)
CORS(app, origins=["*"], methods=["GET", "POST", "OPTIONS"])

# Configuration from environment
UPLOAD_FOLDER = os.getenv('UPLOAD_FOLDER', 'temp_uploads')
RESULT_FOLDER = os.getenv('RESULT_FOLDER', 'temp_results')
DEBUG_MODE = os.getenv('DEBUG', '0') == '1'

# Ensure directories exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(RESULT_FOLDER, exist_ok=True)

# Google Cloud Vision client initialization
def initialize_vision_client():
    """Initialize Google Cloud Vision client with proper authentication"""
    try:
        # Option 1: Use service account credentials file
        credentials_path = os.getenv('GOOGLE_APPLICATION_CREDENTIALS')
        if credentials_path and os.path.exists(credentials_path):
            os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = credentials_path
            print(f"Using service account credentials from: {credentials_path}")
            return vision.ImageAnnotatorClient()
        
        # Option 2: Use API key
        api_key = os.getenv('GOOGLE_CLOUD_VISION_API_KEY')
        if api_key:
            print("Using API key for Vision API")
            # For API key, we'll use direct HTTP requests instead of the client library
            # This is because the client library doesn't support API keys directly
            return "API_KEY_MODE"
        
        # Option 3: Default authentication (ADC)
        print("Using default authentication (Application Default Credentials)")
        return vision.ImageAnnotatorClient()
    
    except Exception as e:
        print(f"Error initializing Vision client: {e}")
        print("Make sure to set GOOGLE_APPLICATION_CREDENTIALS or GOOGLE_CLOUD_VISION_API_KEY")
        return None

vision_client = initialize_vision_client()

class FieldMapping:
    def __init__(self, field_name: str, excel_cell: Dict):
        self.field_name = field_name
        self.excel_cell = excel_cell  # {'sheet': str, 'row': int, 'column': int}

def load_field_mappings_from_env() -> tuple:
    """Load field mappings from environment variables for both invoice and export certificate"""
    invoice_mappings = []
    export_certificate_mappings = []
    
    # Define invoice field names
    invoice_field_names = [
        'total_amount', 'email', 'company_address', 'address_line1', 'address_line2', 'item_name', 'company_name', 'pod', 'invoice_postal_country', 'invoice_email', 'invoice_phone'
    ]
    
    # Define export certificate field names (prefixed with EC_)
    export_certificate_field_names = [
        'chassis_number', 'gasoline', 'seat_number', 'weight', 'engine_capacity', 'length', 'width', 'height', 'net_weight', 'year'
    ]
    
    print("Loading invoice field mappings from environment...")
    
    # Load invoice mappings
    for field_name in invoice_field_names:
        # Get Excel cell mapping
        cell_key = f"{field_name.upper()}_CELL"
        cell_value = os.getenv(cell_key)
        
        print(f"Invoice field '{field_name}': cell='{cell_value}'")
        
        if cell_value:
            try:
                # Parse Excel cell: "Sheet1:2:2" -> sheet, row, column
                cell_value_clean = cell_value.split('#')[0].strip()
                parts = cell_value_clean.split(':')
                if len(parts) >= 3:
                    sheet_name = ':'.join(parts[:-2])
                    row_str = parts[-2].strip()
                    col_str = parts[-1].strip()
                else:
                    sheet_name, row_str, col_str = cell_value_clean.split(':')
                
                excel_cell = {
                    'sheet': sheet_name.strip(),
                    'row': int(row_str),
                    'column': int(col_str)
                }
                
                mapping = FieldMapping(field_name, excel_cell)
                invoice_mappings.append(mapping)
                print(f"Successfully loaded invoice mapping for '{field_name}': {excel_cell}")
                
            except (ValueError, IndexError) as e:
                print(f"Error parsing invoice field mapping for {field_name}: {e}")
                continue
    
    print("Loading export certificate field mappings from environment...")
    
    # Load export certificate mappings
    for field_name in export_certificate_field_names:
        # Get Excel cell mapping (prefixed with EC_)
        cell_key = f"EC_{field_name.upper()}_CELL"
        cell_value = os.getenv(cell_key)
        
        print(f"Export certificate field '{field_name}': cell='{cell_value}'")
        
        if cell_value:
            try:
                # Parse Excel cell: "Sheet1:2:2" -> sheet, row, column
                cell_value_clean = cell_value.split('#')[0].strip()
                parts = cell_value_clean.split(':')
                if len(parts) >= 3:
                    sheet_name = ':'.join(parts[:-2])
                    row_str = parts[-2].strip()
                    col_str = parts[-1].strip()
                else:
                    sheet_name, row_str, col_str = cell_value_clean.split(':')
                
                excel_cell = {
                    'sheet': sheet_name.strip(),
                    'row': int(row_str),
                    'column': int(col_str)
                }
                
                # Prefix field name to distinguish from invoice fields
                mapping = FieldMapping(f"ec_{field_name}", excel_cell)
                export_certificate_mappings.append(mapping)
                print(f"Successfully loaded export certificate mapping for '{field_name}': {excel_cell}")
                
            except (ValueError, IndexError) as e:
                print(f"Error parsing export certificate field mapping for {field_name}: {e}")
                continue
    
    print(f"Total invoice mappings loaded: {len(invoice_mappings)}")
    print(f"Total export certificate mappings loaded: {len(export_certificate_mappings)}")
    return invoice_mappings, export_certificate_mappings

# Initialize field mappings as None - will be loaded lazily
INVOICE_FIELD_MAPPINGS = None
EXPORT_CERTIFICATE_FIELD_MAPPINGS = None

def get_field_mappings():
    """Lazily load field mappings to avoid startup issues"""
    global INVOICE_FIELD_MAPPINGS, EXPORT_CERTIFICATE_FIELD_MAPPINGS
    
    if INVOICE_FIELD_MAPPINGS is None or EXPORT_CERTIFICATE_FIELD_MAPPINGS is None:
        print("Loading field mappings...")
        try:
            INVOICE_FIELD_MAPPINGS, EXPORT_CERTIFICATE_FIELD_MAPPINGS = load_field_mappings_from_env()
            
            print(f"Field mappings loaded successfully: {len(INVOICE_FIELD_MAPPINGS)} invoice, {len(EXPORT_CERTIFICATE_FIELD_MAPPINGS)} export certificate")
            
        except Exception as e:
            print(f"Error loading field mappings: {e}")
            # Use empty defaults if loading fails
            INVOICE_FIELD_MAPPINGS = []
            EXPORT_CERTIFICATE_FIELD_MAPPINGS = []
    
    return INVOICE_FIELD_MAPPINGS, EXPORT_CERTIFICATE_FIELD_MAPPINGS

def pdf_to_images(pdf_file) -> List[Image.Image]:
    """Convert PDF to images for OCR processing"""
    try:
        # Save PDF to temporary file
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp_pdf:
            pdf_file.save(tmp_pdf.name)
            
            # Convert PDF to images
            images = pdf2image.convert_from_path(tmp_pdf.name, dpi=300, first_page=1, last_page=1)
            
            # Clean up
            os.unlink(tmp_pdf.name)
            
            return images
    except Exception as e:
        print(f"Error converting PDF to images: {e}")
        return []

def process_image_with_vision(image: Image.Image) -> str:
    """Process image with Google Cloud Vision OCR and return the full raw text."""
    try:
        # Convert PIL Image to bytes
        img_byte_arr = io.BytesIO()
        image.save(img_byte_arr, format='PNG')
        img_byte_arr.seek(0)
        image_content = img_byte_arr.getvalue()
        
        # Handle different authentication modes
        if vision_client == "API_KEY_MODE":
            # Use direct HTTP request with API key
            api_key = os.getenv('GOOGLE_CLOUD_VISION_API_KEY')
            url = f"https://vision.googleapis.com/v1/images:annotate?key={api_key}"
            
            request_body = {
                "requests": [
                    {
                        "image": {
                            "content": base64.b64encode(image_content).decode('utf-8')
                        },
                        "features": [
                            {
                                "type": "DOCUMENT_TEXT_DETECTION", # Use document text detection for better structure
                                "maxResults": 1
                            }
                        ],
                        "imageContext": {
                            "languageHints": ["en", "ja"]
                        }
                    }
                ]
            }
            
            import requests
            response = requests.post(url, json=request_body)
            
            if response.status_code != 200:
                raise Exception(f"Vision API HTTP error: {response.status_code} - {response.text}")
            
            result = response.json()
            
            # Save complete Vision API response to file for debugging if DEBUG_MODE
            if DEBUG_MODE:
                with open("vision_response_debug.json", "w") as f:
                    json.dump(result, f, indent=2)
                print(f"Complete Vision API response saved to vision_response_debug.json")
            
            if 'error' in result:
                raise Exception(f"Vision API error: {result['error']}")
            
            # Process the response
            if 'responses' in result and result['responses'] and 'fullTextAnnotation' in result['responses'][0]:
                full_text = result['responses'][0]['fullTextAnnotation']['text']
                print(f"Vision API (API Key) processed successfully. Full text extracted.")
                return full_text
            else:
                print("Vision API (API Key) did not return full text annotation.")
                return ""
            
        else:
            # Use client library
            vision_image = vision.Image(content=image_content)
            
            # Set up image context for language hints
            image_context = vision.ImageContext(language_hints=["en", "ja"])
            response = vision_client.document_text_detection(image=vision_image, image_context=image_context) # Use document text detection
            
            # Save complete Vision API response to file for debugging if DEBUG_MODE
            if DEBUG_MODE:
                response_dict = {
                    'text_annotations': [
                        {
                            'description': annotation.description,
                            'bounding_poly': {
                                'vertices': [{'x': v.x, 'y': v.y} for v in annotation.bounding_poly.vertices]
                            } if annotation.bounding_poly.vertices else None,
                        } for annotation in response.text_annotations
                    ] if response.text_annotations else [],
                    'error': response.error.message if response.error.message else None
                }
                with open("vision_response_debug.json", "w") as f:
                    json.dump(response_dict, f, indent=2)
                print(f"Complete Vision API response saved to vision_response_debug.json")
            
            if response.error.message:
                raise Exception(f'Vision API error: {response.error.message}')
            
            if response.full_text_annotation:
                print(f"Vision API (Client Library) processed successfully. Full text extracted.")
                return response.full_text_annotation.text
            else:
                print("Vision API (Client Library) did not return full text annotation.")
                return ""
        
    except Exception as e:
        print(f"Error processing image with Vision API: {e}")
        print(f"Vision client status: {vision_client}")
        return ""

def extract_data_with_openai(ocr_text: str, document_type: str, fields_to_extract: List[str]) -> Dict[str, str]:
    """
    Use OpenAI's GPT model to extract structured data from OCR text.

    :param ocr_text: Raw text extracted from a document by an OCR service.
    :param document_type: The type of document (e.g., 'invoice', 'export_certificate').
    :param fields_to_extract: A list of field names to be extracted from the text.
    :return: A dictionary containing the extracted data.
    """
    if not openai_client:
        raise Exception("OpenAI client is not initialized. Cannot proceed with data extraction.")

    prompt = ""
    if document_type == "invoice":
        # A detailed, specialized prompt for invoices to improve accuracy
        special_invoice_fields = {
            'company_name', 'email', 'company_address', 
            'address_line1', 'address_line2', 'invoice_postal_country', 'invoice_email',
            'invoice_phone', 'item_name'
        }
        other_fields = [f for f in fields_to_extract if f not in special_invoice_fields]
        
        prompt = f"""
You are an intelligent data extraction assistant specializing in logistics and shipping documents.
Your task is to analyze the following text from an **invoice PDF** and identify the specified data fields with high accuracy.
The text may be messy, contain OCR errors, or be in Japanese. You must handle these issues gracefully.
If you encounter Japanese text, translate it to English before extracting the data.

**Very Important Instructions:**
1.  This invoice involves two main parties: the **sender/seller** (the one issuing the invoice) and the **recipient/consignee** (the one receiving the goods). You must carefully distinguish between them.
2.  **For example, the sender/seller is likely 'Turboalfa Japan LLC' and the recipient/consignee is likely 'Harlow Jap Autos'.** Use this as a guide to avoid reversing the companies.

**Field Extraction Rules (Sender/Seller):**
*   `company_name`: Extract the name of the **sender/seller** company (e.g., 'Turboalfa Japan LLC').
*   `email`: Extract the email address of the **sender/seller**. This should logically match the `company_name`.
*   `company_address`: Extract the full address of the **sender/seller**.

**Field Extraction Rules (Recipient/Consignee):**
*   `address_line1`: Extract the company name of the **recipient/consignee** (e.g., 'Harlow Jap Autos').
*   `address_line2`: Extract the rest of the **recipient/consignee's** address (street, city, etc.).
*   `invoice_postal_country`: Extract the postal code and country for the **recipient/consignee**, formatted as "POSTAL_CODE Country".
*   `invoice_email`: Extract the email address of the **recipient/consignee**.
*   `invoice_phone`: Extract the recipient's phone number. **Ensure it starts with a '+' if it is a valid international number.**

**Other Field Rules:**
*   `item_name`: Extract the base vehicle name (e.g., 'Honda S660'). **Remove any year, color, or other descriptive text.**
*   For the remaining fields ({', '.join(other_fields)}), extract their values as you find them.

**Data Cleaning:**
*   Clean the extracted data. For example, amounts should be numbers, dates should be standardized, etc.
*   If a field is not found, its value should be an empty string "".

**OCR Text:**
---
{ocr_text}
---

**Required Fields:**
{json.dumps(fields_to_extract)}

Please provide your response as a single, valid JSON object, following all the rules above.
"""
    else:
        # The original, more generic prompt for other document types
        prompt = f"""
    You are an intelligent data extraction assistant specializing in logistics and shipping documents.
    Your task is to analyze the following text, which was extracted from a {document_type} PDF, and identify the specified data fields.
    The text may be messy and contain OCR errors or be in Japanese. You must handle these issues gracefully.
    If you encounter Japanese text, translate it to English before extracting the data.

    **Instructions:**
    1.  Carefully read the OCR text provided below.
    2.  Identify and extract the values for the following fields: {', '.join(fields_to_extract)}.
    3.  Clean the extracted data. For example, amounts should be numbers, dates should be standardized, etc.
    4.  Return the data as a JSON object where the keys are the field names and the values are the extracted information.
    5.  If a field cannot be found, the value should be an empty string "".

    **Field-Specific Rules:**
    *   `engine_capacity`: Provide the engine capacity in cubic centimeters (cc). If the text says '0.65L', convert it to '650'.
    *   `length`, `width`, `height`: These dimensions must be in meters. If a value is provided in centimeters (cm), you must convert it to meters (e.g., '420cm' becomes '4.2').
    *   `weight`: This is the **gross weight** and it should be greater than the `net_weight`.

    **OCR Text:**
    ---
    {ocr_text}
    ---

    **Required Fields:**
    {json.dumps(fields_to_extract)}

    Please provide your response as a single, valid JSON object.
    """

    try:
        print(f"Sending request to OpenAI for {document_type} data extraction...")
        response = openai_client.chat.completions.create(
            model="o4-mini",
            messages=[
                {"role": "system", "content": "You are a data extraction expert for shipping and invoice documents."},
                {"role": "user", "content": prompt}
            ],
            response_format={"type": "json_object"},
        )
        
        extracted_json = response.choices[0].message.content
        print(f"Received response from OpenAI for {document_type}.")
        
        # Print the raw JSON response for debugging
        print(f"--- OpenAI Raw Response for {document_type} ---")
        print(extracted_json)
        print("-------------------------------------------------")
        
        if DEBUG_MODE:
            debug_filename = f"openai_response_{document_type}_{uuid.uuid4().hex[:6]}.json"
            with open(os.path.join(RESULT_FOLDER, debug_filename), "w") as f:
                json.dump(json.loads(extracted_json), f, indent=2)
            print(f"OpenAI response for {document_type} saved to {debug_filename}")

        return json.loads(extracted_json)

    except Exception as e:
        print(f"Error calling OpenAI API: {e}")
        # Return a dictionary with empty values so the process can continue
        return {field: "" for field in fields_to_extract}


def populate_excel_template(template_path, extracted_data: Dict[str, Dict]) -> str:
    """Populate Excel template with extracted data at specific cell coordinates"""
    try:
        # Load the workbook using openpyxl directly from the template path
        workbook = openpyxl.load_workbook(template_path)
        
        # Determine the target worksheet in a robust way
        worksheet = None
        target_sheet_name = None

        # Find the first defined sheet name from the configuration
        for data_dict in extracted_data.values():
            if data_dict.get('excel_cell') and data_dict['excel_cell'].get('sheet'):
                target_sheet_name = data_dict['excel_cell']['sheet']
                break
        
        if not target_sheet_name:
             # Fallback to the active sheet if no sheet is specified in the config
            worksheet = workbook.active
            print("Warning: No sheet name specified in config. Falling back to active sheet.")
        else:
            # Find the sheet with a case-insensitive and space-insensitive match
            for sheet in workbook.sheetnames:
                # Convert sheet name to string to handle cases where it might be a number (e.g., a year)
                if str(sheet).strip().lower() == str(target_sheet_name).strip().lower():
                    worksheet = workbook[sheet]
                    break
        
        if not worksheet:
            # If still no worksheet is found, raise an error instead of creating a new one.
            raise FileNotFoundError(f"Could not find a worksheet named '{target_sheet_name}' in the template. "
                                    f"Available sheets: {workbook.sheetnames}")

        print(f"Populating data into sheet: '{worksheet.title}'")

        # Populate cells with extracted data
        for field_name, data_dict in extracted_data.items():
            if data_dict.get('text') and data_dict.get('excel_cell'):
                excel_cell = data_dict['excel_cell']
                text = data_dict['text']
                
                # Convert text to appropriate Excel data type
                excel_value = convert_to_excel_value(text, field_name)
                
                # Write to specific cell
                row = excel_cell.get('row', 1)
                col = excel_cell.get('column', 1)
                
                # Write the value to the cell
                cell = worksheet.cell(row=row, column=col)
                
                # Handle merged cells - if cell is merged, write to the top-left cell of the merged range
                if cell.coordinate in worksheet.merged_cells:
                    # Find the top-left cell of the merged range
                    for merged_range in worksheet.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            top_left_cell = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                            top_left_cell.value = excel_value
                            print(f"Populated {field_name} = {excel_value} (type: {type(excel_value).__name__}) at merged cell {top_left_cell.coordinate}")
                            break
                else:
                    cell.value = excel_value
                    print(f"Populated {field_name} = {excel_value} (type: {type(excel_value).__name__}) at {worksheet.title}:{row}:{col}")
        
        # Add today's date to cell 6-O
        try:
            # Format today's date as '10-Jul-25'
            today = datetime.now()
            formatted_date = today.strftime('%d-%b-%y')
            
            # Write to cell 6-O (row 6, column 15)
            date_cell = worksheet.cell(row=6, column=15)
            date_cell.value = formatted_date
            print(f"Populated today's date = {formatted_date} at cell 6-O")
            
        except Exception as e:
            print(f"Error adding today's date: {e}")
        
        # Auto-adjust column widths for better readability
        for column_cells in worksheet.columns:
            try:
                max_length = max(len(str(cell.value)) for cell in column_cells if cell.value)
                column_letter = get_column_letter(column_cells[0].column)
                worksheet.column_dimensions[column_letter].width = max_length + 2
            except (ValueError, TypeError):
                # Ignore empty columns or other issues
                pass
        
        # Save the populated workbook
        result_filename = f"populated_{uuid.uuid4().hex}.xlsx"
        result_path = os.path.join(RESULT_FOLDER, result_filename)
        workbook.save(result_path)
        
        return result_filename
        
    except Exception as e:
        print(f"Error populating Excel template: {e}")
        # Create a simple Excel file with extracted data
        simple_data = {field: data.get('text', '') for field, data in extracted_data.items()}
        df = pd.DataFrame([simple_data])
        result_filename = f"simple_{uuid.uuid4().hex}.xlsx"
        result_path = os.path.join(RESULT_FOLDER, result_filename)
        df.to_excel(result_path, index=False)
        return result_filename

def convert_to_excel_value(text: str, field_name: str) -> any:
    """Convert text to appropriate Excel data type based on field name"""
    if not text:
        return text
    
    # Convert to string and remove any leading/trailing whitespace
    text = str(text).strip()
    
    # Fields that should be numbers
    numeric_fields = [
        'total_amount', 'weight', 'net_weight', 'length', 'width', 'height', 
        'engine_capacity', 'seat_number', 'year'
    ]
    
    # Fields that should be text (emails, names, addresses, etc.)
    text_fields = [
        'company_name', 'item_name', 'chassis_number', 'gasoline', 'pod',
        'email', 'invoice_email', 'invoice_phone', 'invoice_postal_country'
    ]
    
    # Check if this field should be numeric
    if any(field in field_name.lower() for field in numeric_fields):
        try:
            # Remove common non-numeric characters for amount fields
            if 'amount' in field_name.lower():
                # Remove currency symbols and commas, keep decimal points
                cleaned_text = text.replace(',', '').replace('¥', '').replace('$', '').replace('€', '').replace('£', '')
                return float(cleaned_text)
            else:
                # For other numeric fields, just convert to float/int
                if '.' in text:
                    return float(text)
                else:
                    return int(text)
        except (ValueError, TypeError):
            # If conversion fails, return as text
            return text
    
    # For text fields, return as is
    return text

@app.route('/', methods=['GET'])
def root():
    """Root endpoint for basic connectivity"""
    return jsonify({
        'message': 'PDF Processor API is running',
        'status': 'ok',
        'timestamp': datetime.now().isoformat()
    })

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    # Load field mappings if not already loaded
    invoice_mappings, export_mappings = get_field_mappings()
    
    return jsonify({
        'status': 'healthy', 
        'timestamp': datetime.now().isoformat(),
        'vision_client_ready': vision_client is not None,
        'openai_client_ready': openai_client is not None,
        'invoice_field_mappings_loaded': len(invoice_mappings),
        'export_certificate_field_mappings_loaded': len(export_mappings),
        'upload_folder': UPLOAD_FOLDER,
        'result_folder': RESULT_FOLDER
    })

@app.route('/api/test-vision', methods=['POST'])
def test_vision():
    """Test endpoint for Vision API"""
    try:
        if vision_client is None:
            return jsonify({
                'success': False,
                'error': 'Vision client not initialized'
            }), 500
        
        # Test with a simple image or return client status
        return jsonify({
            'success': True,
            'message': 'Vision client is ready',
            'client_type': type(vision_client).__name__
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/test-openai', methods=['GET'])
def test_openai():
    """Test endpoint for OpenAI API connectivity"""
    if not openai_client:
        return jsonify({'success': False, 'error': 'OpenAI client not initialized'}), 500
    try:
        # Make a simple test call
        openai_client.models.list()
        return jsonify({'success': True, 'message': 'OpenAI client is configured and responding correctly.'})
    except Exception as e:
        return jsonify({'success': False, 'error': f'Failed to connect to OpenAI API: {str(e)}'}), 500


# Global progress tracking
progress_sessions = {}

def send_progress_update(session_id: str, step: str, percentage: int, message: str):
    """Send progress update to client"""
    if session_id in progress_sessions:
        progress_data = {
            'step': step,
            'percentage': percentage,
            'message': message,
            'timestamp': datetime.now().isoformat()
        }
        progress_sessions[session_id].put(progress_data)

@app.route('/api/progress/<session_id>', methods=['GET'])
def progress_stream(session_id):
    """Server-Sent Events endpoint for real-time progress updates"""
    def generate():
        if session_id not in progress_sessions:
            progress_sessions[session_id] = Queue()
        
        queue = progress_sessions[session_id]
        
        while True:
            try:
                # Wait for progress updates
                progress_data = queue.get(timeout=30)  # 30 second timeout
                
                # Send SSE data
                yield f"data: {json.dumps(progress_data)}\n\n"
                
                # If we reach 100%, close the stream
                if progress_data.get('percentage', 0) >= 100:
                    break
                    
            except Exception as e:
                # Send error or keep-alive
                yield f"data: {json.dumps({'error': str(e)})}\n\n"
                break
    
    return Response(generate(), mimetype='text/event-stream')

@app.route('/api/process-pdf', methods=['POST'])
def process_pdf():
    """Main endpoint for processing PDF with OCR and Excel template"""
    session_id = str(uuid.uuid4())
    progress_sessions[session_id] = Queue()
    
    try:
        send_progress_update(session_id, "init", 0, "Initializing PDF processing...")
        
        # Check if Vision client is initialized
        if vision_client is None:
            send_progress_update(session_id, "error", 0, "Vision API not configured")
            return jsonify({
                'success': False,
                'error': 'Google Cloud Vision API is not properly configured. Please check your credentials.',
                'session_id': session_id
            }), 500

        # Check if OpenAI client is initialized
        if openai_client is None:
            send_progress_update(session_id, "error", 0, "OpenAI API not configured")
            return jsonify({
                'success': False,
                'error': 'OpenAI API is not properly configured. Please check your OPENAI_API_KEY.',
                'session_id': session_id
            }), 500
        
        send_progress_update(session_id, "validate", 5, "Validating uploaded files...")
        
        # Validate request - both files are now required
        if 'invoice_pdf_file' not in request.files:
            send_progress_update(session_id, "error", 0, "Invoice PDF file is required")
            return jsonify({
                'success': False,
                'error': 'Invoice PDF file is required',
                'session_id': session_id
            }), 400
        
        if 'export_certificate_pdf_file' not in request.files:
            send_progress_update(session_id, "error", 0, "Export certificate PDF file is required")
            return jsonify({
                'success': False,
                'error': 'Export certificate PDF file is required',
                'session_id': session_id
            }), 400
        
        invoice_pdf_file = request.files['invoice_pdf_file']
        export_certificate_pdf_file = request.files['export_certificate_pdf_file']
        
        # Validate file types
        if not invoice_pdf_file.filename.lower().endswith('.pdf'):
            send_progress_update(session_id, "error", 0, "Invalid invoice PDF file format")
            return jsonify({
                'success': False,
                'error': 'Invalid invoice PDF file format',
                'session_id': session_id
            }), 400
        
        if not export_certificate_pdf_file.filename.lower().endswith('.pdf'):
            send_progress_update(session_id, "error", 0, "Invalid export certificate PDF file format")
            return jsonify({
                'success': False,
                'error': 'Invalid export certificate PDF file format',
                'session_id': session_id
            }), 400
        
        # The cleanup call was removed, but the progress update can remain or be removed.
        # For now, I'll remove both to keep things clean.
        # send_progress_update(session_id, "cleanup", 10, "Cleaning up old files...")
        
        # Clean up old files before processing
        # cleanup_old_files() # This function was removed.
        
        send_progress_update(session_id, "template", 15, "Loading Excel template...")
        
        # Use fixed template - handle both local and deployment paths
        template_path = 'template.xlsx'
        if not os.path.exists(template_path):
            # Try with automation directory prefix for deployment
            template_path = os.path.join(os.path.dirname(__file__), 'template.xlsx')
            if not os.path.exists(template_path):
                send_progress_update(session_id, "error", 0, "Template file not found")
                return jsonify({
                    'success': False,
                    'error': 'Template file not found',
                    'session_id': session_id
                }), 500
        
        send_progress_update(session_id, "convert_invoice", 20, "Converting invoice PDF to images...")
        
        # Convert invoice PDF to images
        invoice_images = pdf_to_images(invoice_pdf_file)
        if not invoice_images:
            send_progress_update(session_id, "error", 0, "Failed to convert invoice PDF to images")
            return jsonify({
                'success': False,
                'error': 'Failed to convert invoice PDF to images. Please ensure the PDF is not corrupted and contains at least one page.',
                'session_id': session_id
            }), 500
        
        send_progress_update(session_id, "ocr_invoice", 35, "Processing invoice with OCR...")
        
        # Process invoice first page with OCR
        invoice_ocr_text = process_image_with_vision(invoice_images[0])
        if not invoice_ocr_text:
            send_progress_update(session_id, "error", 0, "Failed to extract text from invoice PDF")
            return jsonify({
                'success': False,
                'error': 'Failed to extract text from invoice PDF. Please check if the PDF contains readable text.',
                'session_id': session_id
            }), 500
        
        send_progress_update(session_id, "convert_export", 50, "Converting export certificate PDF to images...")
        
        # Convert export certificate PDF to images
        export_certificate_images = pdf_to_images(export_certificate_pdf_file)
        if not export_certificate_images:
            send_progress_update(session_id, "error", 0, "Failed to convert export certificate PDF to images")
            return jsonify({
                'success': False,
                'error': 'Failed to convert export certificate PDF to images. Please ensure the PDF is not corrupted and contains at least one page.',
                'session_id': session_id
            }), 500
        
        send_progress_update(session_id, "ocr_export", 65, "Processing export certificate with OCR...")
        
        # Process export certificate first page with OCR
        export_certificate_ocr_text = process_image_with_vision(export_certificate_images[0])
        if not export_certificate_ocr_text:
            send_progress_update(session_id, "error", 0, "Failed to extract text from export certificate PDF")
            return jsonify({
                'success': False,
                'error': 'Failed to extract text from export certificate PDF. Please check if the PDF contains readable text.',
                'session_id': session_id
            }), 500
        
        send_progress_update(session_id, "map_fields", 75, "Mapping OCR results to fields...")
        
        # Get field names from mappings
        invoice_mappings, export_mappings = get_field_mappings()
        invoice_fields_to_extract = [m.field_name for m in invoice_mappings]
        export_fields_to_extract = [m.field_name.replace('ec_', '') for m in export_mappings] # Remove 'ec_' prefix for the prompt

        # Extract data using OpenAI
        send_progress_update(session_id, "extract_invoice_openai", 80, "Extracting invoice data with AI...")
        invoice_extracted_data = extract_data_with_openai(invoice_ocr_text, "invoice", invoice_fields_to_extract)

        send_progress_update(session_id, "extract_export_openai", 85, "Extracting export certificate data with AI...")
        export_certificate_extracted_data = extract_data_with_openai(export_certificate_ocr_text, "export certificate", export_fields_to_extract)

        # Re-associate the extracted data with the correct excel cell mappings
        all_mappings = {m.field_name: m.excel_cell for m in invoice_mappings}
        all_mappings.update({m.field_name: m.excel_cell for m in export_mappings})

        final_extracted_data = {}
        for field, value in invoice_extracted_data.items():
            if field in all_mappings:
                final_extracted_data[field] = {
                    'text': value,
                    'excel_cell': all_mappings[field],
                    'confidence': 0.95  # Confidence from OpenAI is generally high
                }

        for field, value in export_certificate_extracted_data.items():
            ec_field_name = f"ec_{field}"
            if ec_field_name in all_mappings:
                 final_extracted_data[ec_field_name] = {
                    'text': value,
                    'excel_cell': all_mappings[ec_field_name],
                    'confidence': 0.95
                }
        
        send_progress_update(session_id, "populate_excel", 90, "Populating Excel template...")
        
        # Populate Excel template
        result_filename = populate_excel_template(template_path, final_extracted_data)
        
        send_progress_update(session_id, "complete", 100, "Processing completed successfully!")
        
        # Prepare response data for the frontend
        response_data = {}
        response_data.update(invoice_extracted_data)
        response_data.update(export_certificate_extracted_data)

        # Return success response
        return jsonify({
            'success': True,
            'message': 'PDFs processed successfully using AI',
            'extracted_data': response_data,
            'download_url': f'/api/download/{result_filename}',
            'session_id': session_id,
            'mapped_fields': len(final_extracted_data),
            'field_details': final_extracted_data,  # Include full details for debugging
            'files_processed': {
                'invoice': invoice_pdf_file.filename,
                'export_certificate': export_certificate_pdf_file.filename
            }
        })
        
    except Exception as e:
        send_progress_update(session_id, "error", 0, f"Processing failed: {str(e)}")
        print(f"Error processing PDF: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'Processing failed: {str(e)}',
            'session_id': session_id
        }), 500
    finally:
        # Clean up progress session after a delay
        def cleanup_session():
            time.sleep(10)  # Keep session for 10 seconds after completion
            if session_id in progress_sessions:
                del progress_sessions[session_id]
            
            # Force garbage collection to free memory
            import gc
            gc.collect()
        
        threading.Thread(target=cleanup_session, daemon=True).start()

@app.route('/api/download/<filename>', methods=['GET'])
def download_file(filename):
    """Download processed Excel file"""
    try:
        file_path = os.path.join(RESULT_FOLDER, filename)
        
        if not os.path.exists(file_path):
            return jsonify({'error': 'File not found'}), 404
        
        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/cleanup', methods=['POST'])
def cleanup_files():
    """Clean up old temporary files"""
    try:
        current_time = datetime.now()
        cutoff_time = current_time - timedelta(hours=1)  # Clean files older than 1 hour
        
        cleaned_count = 0
        
        # Clean upload folder
        for filename in os.listdir(UPLOAD_FOLDER):
            file_path = os.path.join(UPLOAD_FOLDER, filename)
            if os.path.getmtime(file_path) < cutoff_time.timestamp():
                os.remove(file_path)
                cleaned_count += 1
        
        # Clean result folder
        for filename in os.listdir(RESULT_FOLDER):
            file_path = os.path.join(RESULT_FOLDER, filename)
            if os.path.getmtime(file_path) < cutoff_time.timestamp():
                os.remove(file_path)
                cleaned_count += 1
        
        return jsonify({
            'success': True,
            'cleaned_files': cleaned_count
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    print("Starting PDF Processor API...")
    print("Make sure to set GOOGLE_APPLICATION_CREDENTIALS environment variable")
    print("Make sure to set OPENAI_API_KEY in your .env file")
    
    # Get port from environment variable (for production deployment)
    port = int(os.getenv('PORT', 5001))
    debug = os.getenv('DEBUG', 'False').lower() == 'true'
    
    print(f"Starting on port {port}, debug={debug}")
    print(f"Vision client initialized: {vision_client is not None}")
    print(f"OpenAI client initialized: {openai_client is not None}")
    
    # For production, use gunicorn with longer timeout
    if os.getenv('RENDER'):
        # This will be handled by gunicorn in production
        pass
    else:
        app.run(host='0.0.0.0', port=port, debug=debug)
